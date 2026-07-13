"""Excel parsing — flexible header mapping → canonical CRM fields (port of excelService/columnMapper.js)."""
import io
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

ALIASES = {
    "brandName": ["brand name", "brand", "company", "company name", "name", "account", "account name", "organisation", "organization"],
    "brandId": ["brand id", "brandid", "id", "brand code", "account id", "crm id", "company id"],
    "owner": ["owner", "poc", "poc owner", "sales owner", "assigned to", "rep", "sales rep", "account owner"],
    "agency": ["agency", "agency name", "media agency", "ad agency"],
    "duplicateStatus": ["duplicate status", "duplicate", "is duplicate", "dup", "dup status", "duplicate flag"],
}
CANONICAL = list(ALIASES.keys())


def _norm(v) -> str:
    return " ".join(str("" if v is None else v).strip().lower().replace("_", " ").split())


def map_columns(header_row: List[Any]) -> Dict[str, Any]:
    normalized = [_norm(h) for h in (header_row or [])]
    mapping, matched, missing = {}, [], []
    for field, aliases in ALIASES.items():
        idx = -1
        for alias in aliases:
            if alias in normalized:
                idx = normalized.index(alias)
                break
        mapping[field] = idx
        (matched if idx != -1 else missing).append(field)
    return {"map": mapping, "matched": matched, "missing": missing}


def _find_header(rows: List[List[Any]]) -> Optional[Dict[str, Any]]:
    best = None
    for r in range(min(len(rows), 15)):
        m = map_columns(rows[r])
        if m["map"]["brandName"] >= 0 and (best is None or len(m["matched"]) > best["matched"]):
            best = {"index": r, "matched": len(m["matched"]), "mapping": m}
    return best


def _cell(v) -> str:
    return "" if v is None else str(v).strip()


def extract_from_rows(rows: List[List[Any]]) -> Dict[str, Any]:
    if not rows:
        return {"companies": [], "mapping": {"map": {}, "matched": [], "missing": CANONICAL}, "skipped": 0, "error": "The sheet is empty."}
    header = _find_header(rows)
    if not header:
        return {"companies": [], "mapping": map_columns(rows[0]), "skipped": 0,
                "error": "Could not find a Brand Name column in the first rows."}
    m = header["mapping"]["map"]
    companies, skipped = [], 0
    for r in range(header["index"] + 1, len(rows)):
        row = rows[r]
        if not row:
            continue
        brand = _cell(row[m["brandName"]]) if m["brandName"] < len(row) else ""
        if not brand:
            skipped += 1
            continue
        companies.append({
            "rowIndex": r + 1, "brandName": brand,
            "brandId": _cell(row[m["brandId"]]) if 0 <= m["brandId"] < len(row) else "",
            "owner": _cell(row[m["owner"]]) if 0 <= m["owner"] < len(row) else "",
            "agency": _cell(row[m["agency"]]) if 0 <= m["agency"] < len(row) else "",
            "duplicateStatus": _cell(row[m["duplicateStatus"]]) if 0 <= m["duplicateStatus"] < len(row) else "",
        })
    return {"companies": companies, "mapping": header["mapping"], "skipped": skipped}


def parse_bytes(data: bytes) -> Dict[str, Any]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    result = extract_from_rows(rows)
    result["sheetName"] = wb.sheetnames[0]
    return result
